#!/usr/bin/python
import openpyxl, os, sys

alteris_excel_sheet = openpyxl.load_workbook('Alteris_Import_Sheet.xlsx')
data_sheet = alteris_excel_sheet['Import_Asset']
alteris_headers=['Name','Computer Type']
alteris_col_indice={}
alteris_letters=[]
numbers=['0','1','2','3','4','5','6','7','8','9']
def search_maunfacturing(machine_name):
	with open('manufacturing.txt', 'r') as manufacturing_file:
		for line in manufacturing_file:
			if str(machine_name) in line.strip().split('\t')[:1]:
				return 1;
	
	return 0;

	

def add_to_alteris_sheet():
	for row in data_sheet.iter_rows('A1:AG1'):
		for cell in row:
			if cell.value in alteris_headers:
				temptwo = str(cell).strip('<Cell Import_Asset>').strip('.').strip('1')
				alteris_col_indice[temptwo] = cell.value
				alteris_letters.append(temptwo)
	for row in data_sheet.iter_rows('A1:AG1105'):
		for cell in row:
			if search_maunfacturing(cell.value):
				print cell, cell.value
				tempthree = str(cell).strip('<Cell Import_Asset>').strip('.')[1:]
				row_to_search = str('A'+tempthree+':AG'+tempthree)
				#print row_to_search
				for row in data_sheet.iter_rows(row_to_search):
					for cell in row:
						tempfour = str(cell).strip('<Cell Import_Asset>').strip('.')[0:2]
						for num in numbers:
							if tempfour.endswith(num):
								tempfour = str(cell).strip('<Cell Import_Asset>').strip('.')[0:1]
						if tempfour in alteris_col_indice:
							if alteris_col_indice[tempfour]=='Computer Type':
								cell.value = 'Manufacturing'
								
									
#print(search_maunfacturing("TUA1-W03181"))
									
add_to_alteris_sheet()							
alteris_excel_sheet.save('Alteris_Import_Sheet.xlsx')