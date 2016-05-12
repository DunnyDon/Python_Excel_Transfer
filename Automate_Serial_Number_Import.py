#!/usr/bin/python
import openpyxl, os, sys

alteris_excel_sheet = openpyxl.load_workbook('Alteris_Import_Sheet.xlsx')
#asset_excel_sheet = openpyxl.load_workbook('TUA1 PC Asset List.xlsx')
warranty_sheet_excel = openpyxl.load_workbook('Warranty_Info.xlsx')

#laptop_sheet = asset_excel_sheet['Laptops']
#workstation_sheet = asset_excel_sheet['Workstations']
data_sheet = alteris_excel_sheet['Import_Asset']
warranty_sheet = warranty_sheet_excel['Warranty']


print (data_sheet['A2'].value)
#print(laptop_sheet['A4'].value)
#print(workstation_sheet['A4'].value)
numbers=['0','1','2','3','4','5','6','7','8','9']
warranty_col_indice={}
alteris_col_indice={}
warranty_letters =[]
alteris_letters=[]
storage = {}
warranty_headers= ['Serial Number','Warranty Status','Warranty Start','Warranty End']
alteris_headers = ['Serial Number', 'Warranty Period (Months)','Available For Use','Planned Disposal Date', 'Start Date Warranty', 'End date Warranty', 'Status Warranty']

def get_headers():
	for row in warranty_sheet.iter_rows('A1:U1'):
		for cell in row:
			if cell.value in warranty_headers:
				temp = str(cell).strip('<Cell Warranty>').strip('.').strip('1')
				warranty_col_indice[temp] = cell.value
				warranty_letters.append(temp)
				#print temp, cell.value
			
	#def alteris_headers():
	for row in data_sheet.iter_rows('A1:AG1'):
		for cell in row:
			if cell.value in alteris_headers:
				temptwo = str(cell).strip('<Cell Import_Asset>').strip('.').strip('1')
				alteris_col_indice[temptwo] = cell.value
				alteris_letters.append(temptwo)
				#print temptwo, cell.value
			
	return;
def search_alteris_Sheet(snum):
	for row in data_sheet.iter_rows('A1:AG1105'):
		for cell in row:
			#print cell.value, snum
			if cell.value == snum:
				tempthree = str(cell).strip('<Cell Import_Asset>').strip('.')[1:]
				row_to_search = str('A'+tempthree+':AG'+tempthree)
				print row_to_search
				for row in data_sheet.iter_rows(row_to_search):
					for cell in row:
						tempfour = str(cell).strip('<Cell Import_Asset>').strip('.')[0:2]
						for num in numbers:
							if tempfour.endswith(num):
								tempfour = str(cell).strip('<Cell Import_Asset>').strip('.')[0:1]
						if tempfour in alteris_col_indice:
							for key in warranty_col_indice:
								#print warranty_col_indice[key],  alteris_col_indice[tempfour] 
								if warranty_col_indice[key] == 'Warranty Status' and alteris_col_indice[tempfour]=='Status Warranty':
									if storage[warranty_col_indice[key]].lower() == 'active':
										cell.value = storage[warranty_col_indice[key]]
								elif warranty_col_indice[key] =='Warranty Start':
									startdate=str(storage[warranty_col_indice[key]])[-4:-2]
									startdate+='/'
									startdate+= str(storage[warranty_col_indice[key]])[-2:]
									startdate+='/'
									startdate+=str(storage[warranty_col_indice[key]])[0:4]
									if alteris_col_indice[tempfour]=='Available For Use' or alteris_col_indice[tempfour]=='Start Date Warranty':
										cell.value = startdate
								elif warranty_col_indice[key] =='Warranty End':
									enddate=str(storage[warranty_col_indice[key]])[-4:-2]
									enddate+='/'
									enddate+= str(storage[warranty_col_indice[key]])[-2:]
									enddate+='/'
									enddate+=str(storage[warranty_col_indice[key]])[0:4]
									if alteris_col_indice[tempfour]=='End date Warranty' or alteris_col_indice[tempfour]=='Planned Disposal Date':
										cell.value = enddate
							if alteris_col_indice[tempfour] =='Warranty Period (Months)':
								if int(enddate[-4:]) != 0 or int(startdate[-4:]):
									months_in_years = (int(enddate[-4:]) - int(startdate[-4:]))*12
									months = int(enddate[:2]) - int(startdate[:2])
									days= int(enddate[3:5]) - int(startdate[3:5])
									if days < -10 :
										cell.value = months_in_years + months - 1
									else:
										cell.value = months_in_years + months
							print(alteris_col_indice[tempfour])
							print cell.value
		
	return;
	

get_headers()
for row in warranty_sheet.iter_rows('A2:T1140'):
	for cell in row:
		#print cell
		tempone = str(cell).strip('<Cell Warranty>').strip('.')[:-1]
		if len(tempone) == 2:
			tempone = str(cell).strip('<Cell Warranty>').strip('.')[:-2]
		elif len(tempone) == 3:
			tempone = str(cell).strip('<Cell Warranty>').strip('.')[:-3]
		elif len(tempone) == 4:
			tempone = str(cell).strip('<Cell Warranty>').strip('.')[:-4]
		'''for num in numbers:
			if tempone.endswith(num):
				tempone = str(cell).strip('<Cell Import_Asset>').strip('.')[0:1]
		'''
				#print tempone
		if tempone in warranty_letters:
			#if col_indice[cell.value] in str(cell):
			print warranty_col_indice[tempone], cell.value
			storage[warranty_col_indice[tempone]]= cell.value
	#storage.clear()
	if	storage['Serial Number']!=None :
		search_alteris_Sheet(storage['Serial Number'])
		print '\n'
		'''for key in storage:
			if storage[key]=='':
				tempone = str(cell).strip('<Cell Warranty>').strip('.')[:-1]
				print tempone, cell
				#sys.exit()
			storage[key]='''
		#print tempone, cell
alteris_excel_sheet.save('Alteris_Import_Sheet.xlsx')
#alteris_headers()