#!/usr/bin/python
import openpyxl, os, sys

ad_excel_sheet = openpyxl.load_workbook('AD_DATA.xlsx')
wsus_sheet_excel = openpyxl.load_workbook('Computers Report for TUA1-SV00001.xlsx')

data_sheet = ad_excel_sheet['AD_LIST']
wsus_sheet = wsus_sheet_excel['ComputerStatus']


print (data_sheet['A2'].value)
#print(laptop_sheet['A4'].value)
#print(workstation_sheet['A4'].value)
numbers=['0','1','2','3','4','5','6','7','8','9']
wsus_col_indice={}
ad_col_indice={}
wsus_letters =[]
ad_letters=[]
storage = {}
wsus_headers= ['Name','Needed','Failed','No Status']
ad_headers = ['Name', 'Description']

def get_headers():
	for row in wsus_sheet.iter_rows('A1:H1'):
		for cell in row:
			if cell.value in wsus_headers:
				temp = str(cell).strip('<Cell ComputerStatus>').strip('.').strip('1')
				wsus_col_indice[temp] = cell.value
				wsus_letters.append(temp)
				#print temp, cell.value
			
	#def alteris_headers():
	for row in data_sheet.iter_rows('A1:C1'):
		for cell in row:
			if cell.value in ad_headers:
				temptwo = str(cell).strip('<Cell AD_LIST>').strip('.').strip('1')
				ad_col_indice[temptwo] = cell.value
				ad_letters.append(temptwo)
				#print temptwo, cell.value
			
	return;
	
	
def write_to_log(information):
	#get the current time of and log this also
	try:
		with open('ThePrey3.0.txt', 'a') as logfile:			
			logfile.write(information+"\n")				
	except IOError:
		print("There was an error writing to the LogFile")
		sys.exit(0)
		#exit script if unable to open logfile
	return;
	
	
def search_ad_Sheet(name):
	for row in data_sheet.iter_rows('A1:C1073'):
		for cell in row:
			#print unicode(cell.value).lower(),'a', name
			if unicode(cell.value).lower() in unicode(name):
				tempthree = str(cell).strip('<Cell AD_LIST>').strip('.')[1:]
				row_to_search = str('A'+tempthree+':C'+tempthree)
				for row in data_sheet.iter_rows(row_to_search):
					for cell in row:
						tempfour = str(cell).strip('<Cell AD_LIST>').strip('.')[0:2]
						for num in numbers:
							if tempfour.endswith(num):
								tempfour = str(cell).strip('<Cell AD_LIST>').strip('.')[0:1]
						if tempfour in ad_col_indice:
							for key in wsus_col_indice:
								if wsus_col_indice[key] == 'Name' or wsus_col_indice[key] == 'Description':
									#print ad_col_indice[tempfour],': ', cell.value
									write_to_log(str(ad_col_indice[tempfour])+': '+str(cell.value))

					for key in wsus_col_indice:
						if wsus_col_indice[key] =='Needed':
							write_to_log(wsus_col_indice[key]+': '+ str(storage[wsus_col_indice[key]]))
						elif wsus_col_indice[key] =='Failed':
							write_to_log(wsus_col_indice[key]+': '+ str(storage[wsus_col_indice[key]]))
						elif wsus_col_indice[key] =='No Status':
							write_to_log(wsus_col_indice[key]+': '+ str(storage[wsus_col_indice[key]]))
	return;
	

get_headers()
for row in wsus_sheet.iter_rows('A2:H131'):
	for cell in row:
		tempone = str(cell).strip('<Cell ComputerStatus>').strip('.')[:-1]
		if len(tempone) == 2:
			tempone = str(cell).strip('<Cell ComputerStatus>').strip('.')[:-2]
		elif len(tempone) == 3:
			tempone = str(cell).strip('<Cell ComputerStatus>').strip('.')[:-3]
		elif len(tempone) == 4:
			tempone = str(cell).strip('<Cell ComputerStatus>').strip('.')[:-4]
		'''for num in numbers:
			if tempone.endswith(num):
				tempone = str(cell).strip('<Cell Import_Asset>').strip('.')[0:1]
		'''
		if tempone in wsus_letters:
			#if col_indice[cell.value] in str(cell):
			#print wsus_col_indice[tempone], cell.value
			storage[wsus_col_indice[tempone]]= cell.value
	#storage.clear()
	if	storage['Name']!=None :
		search_ad_Sheet(storage['Name'])
		write_to_log('\n\n')
		'''for key in storage:
			if storage[key]=='':
				tempone = str(cell).strip('<Cell Warranty>').strip('.')[:-1]
				print tempone, cell
				#sys.exit()
			storage[key]='''
		#print tempone, cell
ad_excel_sheet.save('Alteris_Import_Sheet.xlsx')
#alteris_headers()